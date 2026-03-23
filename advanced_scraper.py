"""
╔══════════════════════════════════════════════════════════════════╗
║         WEB SCRAPER PRO — ALL FEATURES EDITION                  ║
║  Features: AI Parse · Scheduler · Change Detection · Proxy      ║
║            Login · Excel/DB Export · 3-Engine Fallback          ║
╚══════════════════════════════════════════════════════════════════╝

INSTALL:
    pip install -r requirements.txt
    playwright install chromium

USAGE:
    python advanced_scraper.py https://srmist.edu.in
    python advanced_scraper.py https://site.com --mode links --output excel
    python advanced_scraper.py https://site.com --schedule daily --notify telegram
    python advanced_scraper.py https://site.com --watch               # change detection
    python advanced_scraper.py https://site.com --proxy-file proxies.txt
    python advanced_scraper.py https://site.com --login user=x pass=y
    python advanced_scraper.py https://site.com --ai-parse            # AI labeling
    python advanced_scraper.py https://site.com --output sqlite
    python advanced_scraper.py https://site.com --playwright          # force real browser
"""

import sys, os, re, json, csv, time, hashlib, sqlite3, argparse, smtplib
import urllib3, threading, schedule as sched_lib
from datetime import datetime
from urllib.parse import urljoin, urlparse
from urllib.robotparser import RobotFileParser
from pathlib import Path
from email.mime.text import MIMEText

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── Optional rich ──────────────────────────────────────────────
try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    console = Console()
    RICH = True
except ImportError:
    RICH = False
    console = None

# ── Optional cloudscraper ──────────────────────────────────────
try:
    import cloudscraper
    HAS_CLOUDSCRAPER = True
except ImportError:
    HAS_CLOUDSCRAPER = False

# ── Optional Playwright ────────────────────────────────────────
try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

# ── Optional openpyxl ──────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# ── Optional requests ──────────────────────────────────────────
import requests
from bs4 import BeautifulSoup


# ══════════════════════════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════════════════════════

def log(msg, style=""):
    ts = datetime.now().strftime("%H:%M:%S")
    if RICH:
        console.print(f"[dim]{ts}[/dim] {msg}", style=style)
    else:
        clean = re.sub(r'\[.*?\]', '', msg)
        print(f"{ts} {clean}")

def banner():
    if RICH:
        console.print(Panel.fit(
            "[bold cyan]Web Scraper Pro[/bold cyan]  [dim]v3.0 — All Features[/dim]\n"
            "[dim]AI · Scheduler · Watch · Proxy · Login · Excel/DB[/dim]",
            border_style="cyan"
        ))
    else:
        print("=" * 60)
        print("  WEB SCRAPER PRO v3.0 — All Features")
        print("=" * 60)


# ══════════════════════════════════════════════════════════════
#  HEADERS
# ══════════════════════════════════════════════════════════════

def get_headers(ua=None):
    return {
        "User-Agent": ua or (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Cache-Control": "max-age=0",
    }


# ══════════════════════════════════════════════════════════════
#  PROXY MANAGER
# ══════════════════════════════════════════════════════════════

class ProxyManager:
    def __init__(self, proxy_file=None, proxy_list=None):
        self.proxies = []
        self.index = 0
        self.lock = threading.Lock()
        if proxy_file and Path(proxy_file).exists():
            with open(proxy_file) as f:
                self.proxies = [l.strip() for l in f if l.strip()]
            log(f"[green]Loaded {len(self.proxies)} proxies from {proxy_file}[/green]")
        elif proxy_list:
            self.proxies = proxy_list

    def get(self):
        if not self.proxies:
            return None
        with self.lock:
            proxy = self.proxies[self.index % len(self.proxies)]
            self.index += 1
        if not proxy.startswith("http"):
            proxy = "http://" + proxy
        return {"http": proxy, "https": proxy}

    def remove_bad(self, proxy_url):
        self.proxies = [p for p in self.proxies if proxy_url not in p]
        log(f"[yellow]Removed bad proxy. {len(self.proxies)} remaining.[/yellow]")

    @property
    def active(self):
        return len(self.proxies) > 0


# ══════════════════════════════════════════════════════════════
#  FETCH ENGINES (3 fallback levels)
# ══════════════════════════════════════════════════════════════

def fetch_with_requests(url, proxy_mgr=None, delay=0.5):
    time.sleep(delay)
    session = requests.Session()
    session.headers.update(get_headers())
    proxies = proxy_mgr.get() if proxy_mgr and proxy_mgr.active else None
    for attempt in range(3):
        try:
            resp = session.get(url, timeout=20, verify=False,
                               allow_redirects=True, proxies=proxies)
            resp.raise_for_status()
            log(f"[green]  ✓ requests OK (HTTP {resp.status_code})[/green]")
            return resp.text
        except requests.exceptions.ProxyError:
            log(f"[yellow]  Proxy error, rotating...[/yellow]")
            if proxy_mgr:
                proxies = proxy_mgr.get()
        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout,
                requests.exceptions.SSLError) as e:
            log(f"[yellow]  Attempt {attempt+1} failed: {type(e).__name__}[/yellow]")
            time.sleep(1.5)
        except requests.exceptions.HTTPError as e:
            code = e.response.status_code
            log(f"[yellow]  HTTP {code} on attempt {attempt+1}[/yellow]")
            if code in (403, 429, 503):
                time.sleep(3)
            else:
                break
    return None


def fetch_with_cloudscraper(url, proxy_mgr=None, delay=0.5):
    if not HAS_CLOUDSCRAPER:
        return None
    time.sleep(delay)
    try:
        scraper = cloudscraper.create_scraper(
            browser={"browser": "chrome", "platform": "windows", "mobile": False}
        )
        proxies = proxy_mgr.get() if proxy_mgr and proxy_mgr.active else None
        resp = scraper.get(url, timeout=20, verify=False, proxies=proxies)
        resp.raise_for_status()
        log(f"[green]  ✓ cloudscraper OK (HTTP {resp.status_code})[/green]")
        return resp.text
    except Exception as e:
        log(f"[yellow]  cloudscraper failed: {e}[/yellow]")
    return None


def fetch_with_playwright(url, delay=0.5, wait_seconds=3):
    if not HAS_PLAYWRIGHT:
        log("[red]  Playwright not installed:[/red]")
        log("[yellow]    pip install playwright && playwright install chromium[/yellow]")
        return None
    time.sleep(delay)
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-setuid-sandbox",
                      "--disable-blink-features=AutomationControlled"]
            )
            ctx = browser.new_context(
                user_agent=get_headers()["User-Agent"],
                viewport={"width": 1280, "height": 800},
                locale="en-US",
                ignore_https_errors=True,
            )
            page = ctx.new_page()
            page.add_init_script(
                "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
            )
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(wait_seconds * 1000)
            page.evaluate("window.scrollTo(0,document.body.scrollHeight/2)")
            page.wait_for_timeout(800)
            html = page.content()
            browser.close()
            log(f"[green]  ✓ Playwright (real browser) OK[/green]")
            return html
    except Exception as e:
        log(f"[red]  Playwright failed: {e}[/red]")
    return None


def smart_fetch(url, proxy_mgr=None, delay=0.5, force_playwright=False):
    """Auto-escalate: requests → cloudscraper → Playwright."""
    if force_playwright:
        return fetch_with_playwright(url, delay), "playwright"

    log(f"[dim]  → requests...[/dim]")
    html = fetch_with_requests(url, proxy_mgr, delay)
    if html:
        soup = BeautifulSoup(html, "lxml")
        if len(soup.get_text(strip=True)) > 300:
            return html, "requests"
        log("[yellow]  Page looks JS-only, escalating...[/yellow]")

    log(f"[dim]  → cloudscraper...[/dim]")
    html = fetch_with_cloudscraper(url, proxy_mgr, delay)
    if html:
        soup = BeautifulSoup(html, "lxml")
        if len(soup.get_text(strip=True)) > 300:
            return html, "cloudscraper"
        log("[yellow]  Still JS-only, escalating to Playwright...[/yellow]")

    log(f"[dim]  → Playwright...[/dim]")
    html = fetch_with_playwright(url, delay)
    return (html, "playwright") if html else (None, None)


# ══════════════════════════════════════════════════════════════
#  LOGIN SUPPORT
# ══════════════════════════════════════════════════════════════

def login_requests(login_url, username, password, session=None):
    """Attempt form-based login and return authenticated session."""
    if session is None:
        session = requests.Session()
    session.headers.update(get_headers())
    try:
        page = session.get(login_url, verify=False, timeout=15)
        soup = BeautifulSoup(page.text, "lxml")
        form = soup.find("form")
        if not form:
            log("[yellow]  No login form found on page.[/yellow]")
            return session
        action = form.get("action", login_url)
        login_post_url = urljoin(login_url, action)
        payload = {}
        for inp in form.find_all("input"):
            name = inp.get("name", "")
            val  = inp.get("value", "")
            if not name:
                continue
            low = name.lower()
            if any(k in low for k in ["user", "email", "login", "name"]):
                payload[name] = username
            elif any(k in low for k in ["pass", "pwd", "secret"]):
                payload[name] = password
            else:
                payload[name] = val
        resp = session.post(login_post_url, data=payload,
                            verify=False, timeout=15, allow_redirects=True)
        if resp.status_code == 200:
            log(f"[green]  ✓ Login submitted (HTTP {resp.status_code})[/green]")
        else:
            log(f"[yellow]  Login response: HTTP {resp.status_code}[/yellow]")
    except Exception as e:
        log(f"[red]  Login failed: {e}[/red]")
    return session


def login_playwright(login_url, username, password):
    """Browser-based login via Playwright."""
    if not HAS_PLAYWRIGHT:
        log("[red]  Playwright required for browser login.[/red]")
        return None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True,
                args=["--no-sandbox", "--disable-setuid-sandbox"])
            ctx = browser.new_context(ignore_https_errors=True)
            page = ctx.new_page()
            page.goto(login_url, timeout=20000)
            # Fill username
            for sel in ['input[type="email"]', 'input[name*="user"]',
                        'input[name*="email"]', 'input[id*="user"]']:
                try:
                    page.fill(sel, username, timeout=2000)
                    break
                except Exception:
                    pass
            # Fill password
            for sel in ['input[type="password"]']:
                try:
                    page.fill(sel, password, timeout=2000)
                    break
                except Exception:
                    pass
            # Submit
            for sel in ['button[type="submit"]', 'input[type="submit"]',
                        'button:has-text("Login")', 'button:has-text("Sign in")']:
                try:
                    page.click(sel, timeout=2000)
                    break
                except Exception:
                    pass
            page.wait_for_timeout(2000)
            cookies = ctx.cookies()
            browser.close()
            log(f"[green]  ✓ Browser login complete, {len(cookies)} cookies captured[/green]")
            return cookies
    except Exception as e:
        log(f"[red]  Browser login failed: {e}[/red]")
    return None


# ══════════════════════════════════════════════════════════════
#  EXTRACTORS
# ══════════════════════════════════════════════════════════════

def extract_text(soup):
    seen, out = set(), []
    for tag in ["p","h1","h2","h3","h4","li","td","th"]:
        for el in soup.find_all(tag):
            t = el.get_text(strip=True)
            if t and len(t) > 15 and t not in seen:
                seen.add(t); out.append({"tag": tag, "text": t})
    return out

def extract_links(soup, base):
    seen, out = set(), []
    for a in soup.find_all("a", href=True):
        h = a["href"].strip()
        if not h or h.startswith(("#","javascript:","mailto:","tel:")):
            continue
        full = urljoin(base, h)
        if full not in seen:
            seen.add(full)
            out.append({"text": a.get_text(strip=True) or "(no text)",
                        "url": full,
                        "external": urlparse(full).netloc != urlparse(base).netloc})
    return out

def extract_images(soup, base):
    out = []
    for img in soup.find_all("img"):
        src = img.get("src") or img.get("data-src","")
        if src:
            out.append({"src": urljoin(base, src),
                        "alt": img.get("alt",""),
                        "title": img.get("title","")})
    return out

def extract_tables(soup):
    out = []
    for table in soup.find_all("table"):
        headers = [th.get_text(strip=True) for th in table.find_all("th")]
        for tr in table.find_all("tr"):
            cells = [td.get_text(strip=True) for td in tr.find_all(["td","th"])]
            if cells and any(cells):
                if headers and len(cells) == len(headers):
                    out.append(dict(zip(headers, cells)))
                else:
                    out.append({"col_"+str(i): v for i,v in enumerate(cells)})
    return out

def extract_prices(soup):
    pat = re.compile(
        r'[\$£€¥₹₨]\s?\d[\d,]*(?:\.\d{1,2})?'
        r'|\d[\d,]+(?:\.\d{2})?\s?(?:INR|USD|EUR|GBP|Rs)'
    )
    seen, out = set(), []
    for el in soup.find_all(text=True):
        for p in pat.findall(el):
            if p not in seen:
                seen.add(p)
                out.append({"price": p.strip(),
                            "context": el.parent.get_text(strip=True)[:100]})
    return out

def extract_emails(soup):
    return list(set(re.findall(
        r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b',
        soup.get_text()
    )))

def extract_phones(soup):
    return list(set(re.findall(
        r'(?:\+91[\s\-]?)?[6-9]\d{9}'
        r'|(?:\+1[\s\-]?)?\(?\d{3}\)?[\s\-]\d{3}[\s\-]\d{4}',
        soup.get_text()
    )))

def extract_custom(soup, selectors):
    out = []
    if not selectors:
        return out
    anchor_els = soup.select(list(selectors.values())[0])
    for i in range(len(anchor_els)):
        rec = {}
        for field, sel in selectors.items():
            els = soup.select(sel)
            rec[field] = els[i].get_text(strip=True) if i < len(els) else ""
        out.append(rec)
    return out

def autodetect(soup, base):
    title = soup.find("title")
    meta  = soup.find("meta", attrs={"name": "description"})
    data  = {
        "url": base,
        "page_title": title.get_text(strip=True) if title else "",
        "meta_description": meta.get("content","") if meta else "",
        "headings": {
            "h1": [h.get_text(strip=True) for h in soup.find_all("h1")][:10],
            "h2": [h.get_text(strip=True) for h in soup.find_all("h2")][:15],
        },
        "paragraphs": [p.get_text(strip=True) for p in soup.find_all("p")
                       if len(p.get_text(strip=True)) > 30][:20],
        "links":  extract_links(soup, base)[:30],
        "images": extract_images(soup, base)[:20],
        "emails": extract_emails(soup),
        "phones": extract_phones(soup),
    }
    tables = extract_tables(soup)
    if tables: data["tables"] = tables[:50]
    prices = extract_prices(soup)
    if prices: data["prices"] = prices[:30]
    return data


# ══════════════════════════════════════════════════════════════
#  AI PARSING
# ══════════════════════════════════════════════════════════════

def ai_parse(data, api_key=None):
    """Use OpenAI/Claude to intelligently label and summarize scraped data."""
    if not api_key:
        api_key = os.environ.get("OPENAI_API_KEY") or os.environ.get("AI_API_KEY")
    if not api_key:
        log("[yellow]  AI parse skipped — set OPENAI_API_KEY env var or --ai-key[/yellow]")
        return data
    try:
        import openai
        client = openai.OpenAI(api_key=api_key)
        snippet = json.dumps(data, ensure_ascii=False)[:3000]
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{
                "role": "user",
                "content": (
                    "You are a data extraction expert. "
                    "Analyze this scraped web data and return a clean JSON with:\n"
                    "- summary: one sentence about the page\n"
                    "- key_data: the most important extracted fields\n"
                    "- data_type: what kind of page this is\n"
                    "- entities: people, organizations, locations found\n\n"
                    f"Data:\n{snippet}\n\n"
                    "Respond ONLY with valid JSON."
                )
            }],
            max_tokens=800,
        )
        ai_result = json.loads(resp.choices[0].message.content)
        data["ai_analysis"] = ai_result
        log("[green]  ✓ AI parse complete[/green]")
    except ImportError:
        log("[yellow]  openai package not installed: pip install openai[/yellow]")
    except Exception as e:
        log(f"[yellow]  AI parse error: {e}[/yellow]")
    return data


# ══════════════════════════════════════════════════════════════
#  CHANGE DETECTION
# ══════════════════════════════════════════════════════════════

HISTORY_FILE = Path("scrape_history.json")

def load_history():
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE) as f:
            return json.load(f)
    return {}

def save_history(history):
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2)

def get_content_hash(data):
    return hashlib.md5(
        json.dumps(data, sort_keys=True, ensure_ascii=False).encode()
    ).hexdigest()

def detect_changes(url, new_data):
    """Compare new data against last scrape. Returns (changed, diff_summary)."""
    history = load_history()
    new_hash = get_content_hash(new_data)
    key = hashlib.md5(url.encode()).hexdigest()

    if key not in history:
        history[key] = {
            "url": url,
            "hash": new_hash,
            "last_scraped": datetime.now().isoformat(),
            "scrape_count": 1,
        }
        save_history(history)
        log("[dim]  First scrape — baseline saved.[/dim]")
        return False, "First scrape — no previous data to compare."

    old_hash = history[key]["hash"]
    if old_hash == new_hash:
        history[key]["last_scraped"] = datetime.now().isoformat()
        history[key]["scrape_count"] = history[key].get("scrape_count", 0) + 1
        save_history(history)
        log("[dim]  No changes detected.[/dim]")
        return False, "No changes detected."

    # Content changed
    old_count = history[key].get("scrape_count", 0)
    diff = (
        f"CHANGE DETECTED on {url}\n"
        f"Previous hash: {old_hash}\n"
        f"New hash:      {new_hash}\n"
        f"Scrape #:      {old_count + 1}\n"
        f"Time:          {datetime.now().isoformat()}"
    )
    history[key] = {
        "url": url,
        "hash": new_hash,
        "last_scraped": datetime.now().isoformat(),
        "scrape_count": old_count + 1,
        "last_change": datetime.now().isoformat(),
    }
    save_history(history)
    log(f"[bold yellow]  ⚡ CHANGE DETECTED on {url}[/bold yellow]")
    return True, diff


# ══════════════════════════════════════════════════════════════
#  NOTIFICATIONS
# ══════════════════════════════════════════════════════════════

def notify_email(subject, body, to_email, smtp_host="smtp.gmail.com",
                 smtp_port=587, from_email=None, from_pass=None):
    from_email = from_email or os.environ.get("NOTIFY_EMAIL")
    from_pass  = from_pass  or os.environ.get("NOTIFY_PASS")
    if not from_email or not from_pass:
        log("[yellow]  Email notify skipped — set NOTIFY_EMAIL and NOTIFY_PASS env vars[/yellow]")
        return
    try:
        msg = MIMEText(body)
        msg["Subject"] = subject
        msg["From"] = from_email
        msg["To"] = to_email
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(from_email, from_pass)
            server.sendmail(from_email, to_email, msg.as_string())
        log(f"[green]  ✓ Email sent to {to_email}[/green]")
    except Exception as e:
        log(f"[red]  Email failed: {e}[/red]")

def notify_telegram(message, bot_token=None, chat_id=None):
    bot_token = bot_token or os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id   = chat_id   or os.environ.get("TELEGRAM_CHAT_ID")
    if not bot_token or not chat_id:
        log("[yellow]  Telegram notify skipped — set TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID[/yellow]")
        return
    try:
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        requests.post(url, data={"chat_id": chat_id, "text": message}, timeout=10)
        log("[green]  ✓ Telegram notification sent[/green]")
    except Exception as e:
        log(f"[red]  Telegram failed: {e}[/red]")


# ══════════════════════════════════════════════════════════════
#  EXPORT — JSON / CSV / EXCEL / SQLITE
# ══════════════════════════════════════════════════════════════

def save_json(data, filename):
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    log(f"[bold green]  → JSON saved: {filename}[/bold green]")

def save_csv(data, filename):
    if not data:
        log("[yellow]  No data for CSV.[/yellow]"); return
    flat = data if isinstance(data[0], dict) else [{"value": str(d)} for d in data]
    keys = list(flat[0].keys())
    with open(filename, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader(); w.writerows(flat)
    log(f"[bold green]  → CSV saved: {filename}[/bold green]")

def save_excel(data, filename):
    if not HAS_EXCEL:
        log("[yellow]  openpyxl not installed: pip install openpyxl[/yellow]")
        save_csv(data, filename.replace(".xlsx", ".csv")); return
    if not data:
        log("[yellow]  No data for Excel.[/yellow]"); return
    flat = data if isinstance(data[0], dict) else [{"value": str(d)} for d in data]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scraped Data"
    header_fill = PatternFill("solid", fgColor="1D9E75")
    header_font = Font(bold=True, color="FFFFFF")
    keys = list(flat[0].keys())
    for col, key in enumerate(keys, 1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = max(15, len(key) + 4)
    for row_i, record in enumerate(flat, 2):
        for col, key in enumerate(keys, 1):
            val = record.get(key, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)[:200]
            ws.cell(row=row_i, column=col, value=str(val) if val else "")
    # Meta sheet
    ws_meta = wb.create_sheet("Meta")
    ws_meta["A1"] = "Scraped at"
    ws_meta["B1"] = datetime.now().isoformat()
    ws_meta["A2"] = "Total records"
    ws_meta["B2"] = len(flat)
    wb.save(filename)
    log(f"[bold green]  → Excel saved: {filename}[/bold green]")

def save_sqlite(data, filename, table="scraped_data"):
    if not data:
        log("[yellow]  No data for SQLite.[/yellow]"); return
    flat = data if isinstance(data[0], dict) else [{"value": str(d)} for d in data]
    keys = list(flat[0].keys())
    conn = sqlite3.connect(filename)
    cur  = conn.cursor()
    safe_keys = [re.sub(r'\W+','_', k) for k in keys]
    cols = ", ".join(f'"{k}" TEXT' for k in safe_keys)
    cur.execute(f'CREATE TABLE IF NOT EXISTS "{table}" (id INTEGER PRIMARY KEY AUTOINCREMENT, scraped_at TEXT, {cols})')
    ts = datetime.now().isoformat()
    for rec in flat:
        vals = [ts] + [str(rec.get(k,""))[:500] if rec.get(k,"") else "" for k in keys]
        placeholders = ",".join(["?"] * len(vals))
        cur.execute(f'INSERT INTO "{table}" (scraped_at,{",".join(chr(34)+k+chr(34) for k in safe_keys)}) VALUES ({placeholders})', vals)
    conn.commit(); conn.close()
    log(f"[bold green]  → SQLite saved: {filename} (table: {table})[/bold green]")

def save_markdown(data, filename):
    lines = [f"# Scraped Data — {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"]
    def fmt(v, d=0):
        ind = "  "*d
        if isinstance(v, dict):
            return "\n".join(f"{ind}- **{k}**: {fmt(vv,d+1)}" for k,vv in v.items())
        if isinstance(v, list):
            return "\n".join(f"{ind}- {fmt(i,d+1)}" for i in v[:20])
        return str(v)
    if isinstance(data, dict):
        for k,v in data.items():
            lines.append(f"## {k}\n{fmt(v)}\n\n")
    elif isinstance(data, list):
        for item in data[:50]:
            lines.append(f"- {fmt(item)}\n")
    with open(filename,"w",encoding="utf-8") as f:
        f.writelines(lines)
    log(f"[bold green]  → Markdown saved: {filename}[/bold green]")


# ══════════════════════════════════════════════════════════════
#  PAGINATION
# ══════════════════════════════════════════════════════════════

def find_next_page(soup, base):
    for text in ["next","next »","›","»","next page","→"]:
        a = soup.find("a", string=lambda s: s and text.lower() in s.lower())
        if a and a.get("href"):
            return urljoin(base, a["href"])
    rel = soup.find("a", rel="next")
    if rel and rel.get("href"):
        return urljoin(base, rel["href"])
    return None


# ══════════════════════════════════════════════════════════════
#  DISPLAY
# ══════════════════════════════════════════════════════════════

def display(data, mode):
    if not RICH:
        print(json.dumps(data, indent=2, ensure_ascii=False)[:3000])
        return
    if mode == "links" and isinstance(data, list):
        t = Table(title="Links", show_lines=True, max_rows=20)
        t.add_column("Text", style="cyan", max_width=35)
        t.add_column("URL",  style="blue", max_width=55)
        t.add_column("Ext",  style="yellow", max_width=5)
        for r in data[:20]:
            t.add_row(r.get("text","")[:35], r.get("url",""), "yes" if r.get("external") else "no")
        console.print(t)
    elif mode in ("tables","custom") and isinstance(data,list) and data:
        first = data[0] if isinstance(data[0],dict) else {}
        if first:
            t = Table(title="Data", show_lines=True, max_rows=20)
            for col in list(first.keys())[:8]:
                t.add_column(col, style="cyan", max_width=28)
            for row in data[:20]:
                t.add_row(*[str(row.get(c,""))[:28] for c in list(first.keys())[:8]])
            console.print(t)
    else:
        console.print_json(json.dumps(data, ensure_ascii=False)[:4000])


# ══════════════════════════════════════════════════════════════
#  CORE SCRAPE
# ══════════════════════════════════════════════════════════════

def scrape(
    url,
    mode="auto",
    selectors=None,
    max_pages=1,
    delay=0.8,
    proxy_mgr=None,
    force_playwright=False,
    login_creds=None,
    ai_parse_flag=False,
    ai_key=None,
    watch=False,
    notify=None,
    notify_target=None,
    output="json",
    output_file=None,
):
    banner()
    log(f"[bold]URL   :[/bold] {url}")
    log(f"[bold]Mode  :[/bold] {mode}  |  Pages: {max_pages}  |  Delay: {delay}s")
    if proxy_mgr and proxy_mgr.active:
        log(f"[bold]Proxy :[/bold] {len(proxy_mgr.proxies)} proxies loaded")
    if login_creds:
        log(f"[bold]Login :[/bold] {login_creds.get('username','?')}")

    # ── Login ──
    session = None
    if login_creds:
        log("\n[cyan]── Logging in...[/cyan]")
        session = login_requests(
            login_creds.get("login_url", url),
            login_creds["username"],
            login_creds["password"],
        )

    all_data, current_url, page_num, methods = [], url, 0, set()

    while current_url and page_num < max_pages:
        page_num += 1
        log(f"\n[cyan]── Page {page_num}: {current_url}[/cyan]")

        html, method = smart_fetch(current_url, proxy_mgr, delay, force_playwright)
        if not html:
            log(f"[red]  ✗ All engines failed for {current_url}[/red]")
            break
        methods.add(method)
        soup = BeautifulSoup(html, "lxml")

        if   mode == "auto":    data = autodetect(soup, current_url); all_data.append(data)
        elif mode == "text":    all_data.extend(extract_text(soup))
        elif mode == "links":   all_data.extend(extract_links(soup, current_url))
        elif mode == "images":  all_data.extend(extract_images(soup, current_url))
        elif mode == "tables":  all_data.extend(extract_tables(soup))
        elif mode == "prices":  all_data.extend(extract_prices(soup))
        elif mode == "emails":  all_data.extend([{"email": e} for e in extract_emails(soup)])
        elif mode == "phones":  all_data.extend([{"phone": p} for p in extract_phones(soup)])
        elif mode == "custom":
            if not selectors:
                log("[red]  --selectors required for custom mode[/red]"); break
            all_data.extend(extract_custom(soup, selectors))

        log(f"[green]  ✓ {len(all_data)} records total (via {method})[/green]")

        if max_pages > 1:
            nxt = find_next_page(soup, current_url)
            current_url = nxt if nxt and nxt != current_url else None
            if not current_url:
                log("[dim]  No more pages.[/dim]")
        else:
            break

    # ── AI parse ──
    if ai_parse_flag and all_data:
        log("\n[cyan]── AI analysis...[/cyan]")
        if isinstance(all_data, list) and len(all_data) > 0:
            all_data[0] = ai_parse(all_data[0], ai_key)
        else:
            all_data = ai_parse(all_data, ai_key)

    # ── Change detection ──
    if watch and all_data:
        log("\n[cyan]── Change detection...[/cyan]")
        changed, diff = detect_changes(url, all_data)
        if changed:
            if notify == "email" and notify_target:
                notify_email(f"[Scraper] Change on {url}", diff, notify_target)
            elif notify == "telegram":
                notify_telegram(diff)

    # ── Summary ──
    log(f"\n[bold]── Summary ──────────────────────────[/bold]")
    log(f"  Records : {len(all_data)}")
    log(f"  Pages   : {page_num}")
    log(f"  Engine  : {', '.join(methods) if methods else 'none'}")

    if not all_data:
        log("[red]  No data. Try --playwright.[/red]")
        return []

    display(all_data, mode)

    # ── Save ──
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    domain = urlparse(url).netloc.replace(".", "_")
    if not output_file:
        ext_map = {"json":"json","csv":"csv","excel":"xlsx",
                   "sqlite":"db","markdown":"md","md":"md"}
        ext = ext_map.get(output, "json")
        output_file = f"scraped_{domain}_{ts}.{ext}"

    if   output == "csv":      save_csv(all_data, output_file)
    elif output == "excel":    save_excel(all_data, output_file)
    elif output == "sqlite":   save_sqlite(all_data, output_file)
    elif output in ("md","markdown"): save_markdown(all_data, output_file)
    else:                      save_json(all_data, output_file)

    log(f"\n[bold green]✓ Done! {len(all_data)} records → {output_file}[/bold green]\n")
    return all_data


# ══════════════════════════════════════════════════════════════
#  SCHEDULER
# ══════════════════════════════════════════════════════════════

def run_scheduled(scrape_args, interval):
    """Run scrape on a schedule. interval = 'hourly'|'daily'|'weekly'|'N' (minutes)."""
    log(f"[cyan]Scheduler started — interval: {interval}[/cyan]")
    log("[dim]Press Ctrl+C to stop.[/dim]\n")

    def job():
        log(f"[cyan]── Scheduled run: {datetime.now().isoformat()}[/cyan]")
        scrape(**scrape_args)

    job()  # run immediately first

    if interval == "hourly":
        sched_lib.every().hour.do(job)
    elif interval == "daily":
        sched_lib.every().day.at("09:00").do(job)
    elif interval == "weekly":
        sched_lib.every().week.do(job)
    else:
        try:
            mins = int(interval)
            sched_lib.every(mins).minutes.do(job)
        except ValueError:
            log(f"[red]Unknown interval '{interval}'. Use hourly/daily/weekly/N.[/red]")
            return

    try:
        while True:
            sched_lib.run_pending()
            time.sleep(30)
    except KeyboardInterrupt:
        log("\n[yellow]Scheduler stopped.[/yellow]")


# ══════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(
        description="Web Scraper Pro v3.0 — All Features",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog="""
EXAMPLES:
  python advanced_scraper.py https://srmist.edu.in
  python advanced_scraper.py https://srmist.edu.in --mode emails
  python advanced_scraper.py https://srmist.edu.in --mode tables --output excel
  python advanced_scraper.py https://srmist.edu.in --mode links --output sqlite
  python advanced_scraper.py https://srmist.edu.in --playwright
  python advanced_scraper.py https://srmist.edu.in --watch --notify telegram
  python advanced_scraper.py https://srmist.edu.in --schedule daily
  python advanced_scraper.py https://srmist.edu.in --proxy-file proxies.txt
  python advanced_scraper.py https://site.com/login --login user=admin pass=secret
  python advanced_scraper.py https://srmist.edu.in --ai-parse --ai-key YOUR_KEY
  python advanced_scraper.py https://srmist.edu.in --pages 5 --delay 1 --output csv
        """
    )
    p.add_argument("url")
    p.add_argument("--mode", default="auto",
                   choices=["auto","text","links","images","tables",
                            "prices","emails","phones","custom"])
    p.add_argument("--selectors", nargs="+", metavar="field=selector")
    p.add_argument("--pages",    type=int,   default=1)
    p.add_argument("--delay",    type=float, default=0.8)
    p.add_argument("--playwright", action="store_true")
    p.add_argument("--proxy-file", help="Path to proxies.txt (one proxy per line)")
    p.add_argument("--login",    nargs="+",  metavar="key=value",
                   help="Login creds: user=X pass=Y login_url=https://...")
    p.add_argument("--ai-parse", action="store_true")
    p.add_argument("--ai-key",   help="OpenAI API key (or set OPENAI_API_KEY env var)")
    p.add_argument("--watch",    action="store_true", help="Enable change detection")
    p.add_argument("--notify",   choices=["email","telegram"],
                   help="Notify on change")
    p.add_argument("--notify-target", help="Email address or Telegram chat ID")
    p.add_argument("--schedule", help="hourly | daily | weekly | N (minutes)")
    p.add_argument("--output",   default="json",
                   choices=["json","csv","excel","sqlite","md","markdown"])
    p.add_argument("--file",     help="Custom output filename")
    args = p.parse_args()

    selectors = None
    if args.selectors:
        selectors = {}
        for s in args.selectors:
            if "=" in s:
                k, v = s.split("=", 1)
                selectors[k.strip()] = v.strip()

    login_creds = None
    if args.login:
        login_creds = {}
        for s in args.login:
            if "=" in s:
                k, v = s.split("=", 1)
                login_creds[k.strip()] = v.strip()
        if "user" in login_creds and "username" not in login_creds:
            login_creds["username"] = login_creds.pop("user")
        if "pass" in login_creds and "password" not in login_creds:
            login_creds["password"] = login_creds.pop("pass")

    proxy_mgr = ProxyManager(proxy_file=args.proxy_file) if args.proxy_file else None

    scrape_args = dict(
        url=args.url,
        mode=args.mode,
        selectors=selectors,
        max_pages=args.pages,
        delay=args.delay,
        proxy_mgr=proxy_mgr,
        force_playwright=args.playwright,
        login_creds=login_creds,
        ai_parse_flag=args.ai_parse,
        ai_key=args.ai_key,
        watch=args.watch,
        notify=args.notify,
        notify_target=args.notify_target,
        output=args.output,
        output_file=args.file,
    )

    if args.schedule:
        run_scheduled(scrape_args, args.schedule)
    else:
        scrape(**scrape_args)


if __name__ == "__main__":
    main()
